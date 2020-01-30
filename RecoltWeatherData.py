import openpyxl as pyxl
# import lightgbm as lgb
import datetime as dt
import pandas as pd
import numpy as np
import argparse
import requests
import calendar
import imaplib
import os.path
import base64
import email
import time
import sys

# from azure.storage.blob import download_blob_from_url, upload_blob_to_url
from vacances_scolaires_france import SchoolHolidayDates
from jours_feries_france.compute import JoursFeries

DATA_FILE = 'Datas/DataGolfEmpty.csv'
DISPLAY_FILE = 'Data/DataDisplay.xlsx'
ADD_DATA = 'Datas/data_after_run_code2.csv'
TOURN_FILE = 'Data/GolfDatesTournois.csv'
TRAIN_DATE_FILE = 'lasttraindate.txt'
NEW_DATA_FILE = 'Datas/DataGolfForecasttest2.csv'

DARKSKY_API_KEY = '0b52a113513d83a0a294d6473aeff133'
DARKSKY_API_KEY2 = '540cb174af287ae97c690ba1a9df5b9b'
DARKSKY_API_KEY3 = '1794bdd2013a58b2e3c6a25708962eef'
DARKSKY_API_KEY4 = '22dcfef0b6342f405fb95cc2b5fe86a3'

BLOB_URL = 'https://displaygolf01.blob.core.windows.net/data/DataDisplay.xlsx'
BLOB_ACCESS_KEY = '1WI/+XcvSm74jZXMM/V3EkeO0EV85x1ZGrOU2sGOIgd+HivVh+YiYJ7ASFMOyhmsVoD35n3sm3JIKM34Ypid4A=='

EMAIL_USER = 'golf.saint-cloud@openvalue.fr'
EMAIL_PASS = 'A08pNyuc8UBUpoibp'

MAX_NB_NEW_DATA_FILES = 100


def check_files(data_file, new_data_file):
    if not os.path.isfile(data_file):
        sys.exit(data_file + ' is not a valid path, use the \'-d\' option to enter a valid path for the data file.')
    if not os.path.isfile(new_data_file):
        sys.exit(
            new_data_file + ' is not a valid path, use the \'--new\' option to enter a valid path for the new data file.')


def parse_date(s):
    return dt.datetime.strptime(s, '%d/%m/%Y').date()


def add_date_part(df, new_cols, drop=True):
    dates = df.index.to_series()

    if not np.issubdtype(dates.dtype, np.datetime64):
        df.index = dates = pd.to_datetime(dates)

    for col in new_cols:
        df[col] = getattr(dates.dt, col)


def shift_cols(df, cols, lags):
    # Check input variables
    assert type(df) is pd.DataFrame, 'ERROR: df must be a DataFrame.'
    assert len(cols) == len(lags), 'ERROR: cols and lags must be arrays of the same length.'

    new_df = df.copy()

    # Create the lagged columns
    for col, lag in zip(cols, lags):
        for i in range(1, lag + 1):
            new_df[col + f'(t - {i})'] = new_df[col].shift(i)

    # Drop rows with Nan values at the beginning
    new_df = new_df.iloc[7:]

    return new_df


def mean_absolute_percentage_error(y_true, y_pred):
    y_true, y_pred = np.array(y_true), np.array(y_pred)
    return np.mean(np.abs((y_true - y_pred) / y_true)) * 100


def prepareData(df):
    # Make 'rain' column bool
    df['rain'] = df['rain'].astype('bool')

    # df['tourn'] = df['tourn'].notnull()
    # Encode bool columns to int
    for col in df.columns:
        if df.dtypes[col] == bool and col != 'open':
            df[col] = df[col].astype('int')

    # Add lagged variables
    cols_to_lag = ['vac', 'tempMoy', 'tempMax', 'tempMin',
                   'wind', 'rain', 'nbPlay']
    nb_lags = [2, 4, 4, 4, 5, 5, 3]
    lagged_df = shift_cols(df, cols_to_lag, nb_lags)

    # Delete rows when the restaurant is closed
    lagged_df = lagged_df[lagged_df['open'] == True]
    lagged_df.drop(columns=['open'], inplace=True)

    # Add time-based variables
    new_cols = ['week', 'day', 'dayofweek', 'dayofyear']
    add_date_part(lagged_df, new_cols)

    return lagged_df


def forecast(df, model, nb_forecast):
    pred_df = df.copy()
    pred_y = np.empty([nb_forecast, 1])

    for i in range(nb_forecast):
        print(f'Forecast #{i}')
        # Lag target
        lagged_df = shift_cols(pred_df, ['nbCov'], [7])

        # Predict for row i
        x = [list(lagged_df.drop(columns='nbCov').values[i])]
        pred_y[i] = model.predict(x)

        # Save prediction for future ones
        pred_df.nbCov.iloc[i + 7] = pred_y[i]

    return pred_y


def get_attribute(weather, key):
    try:
        return weather['daily']['data'][0][key]
    except:
        print('No', key)
        return np.NaN


def check_integrity(df):
    """
        Checks that there's no NaN values, except at the end of the target
        column.
    """
    cols = list(df.columns)

    # Check for NaN values
    for col in cols:
        if col == 'nbCov' and df['nbCov'].isna().any():
            # Get id of first NaN value
            firstna_id = df.loc[df[col].isna()].index[0]

            # Check that all the following rows are also NaN
            if not df.loc[firstna_id:, col].isna().all():
                print(df.loc[firstna_id:, col])
                raise ValueError('Target column has missing values')
        elif df[col].isna().any():
            raise ValueError('Columns ' + col + ' has missing values.')


def add_scale_rain(rain):
    # add a scale on rain
    if rain < 0.2:
        rain = 0
    elif rain >= 0.2 and rain < 0.4:
        rain = 1
    elif rain >= 0.4 and rain < 0.6:
        rain = 2
    elif rain >= 0.6 and rain < 0.8:
        rain = 3
    elif rain >= 0.8:
        rain = 4
    else:
        rain = 0
    return rain


def add_rows(data_df, new_data_df):
    df = data_df.copy()

    shd = SchoolHolidayDates()
    df_tn = pd.read_csv(
        TOURN_FILE,
        index_col=0,
        parse_dates=True,
        date_parser=parse_date,
        sep=';'
    )

    for d in new_data_df.index:

        print('Adding data for date:', d)
        # print(new_data_df.index)

        # Jour feries
        fer = int(d.date() in JoursFeries.for_year(d.year).values())

        # Vacances scolaires
        vac = int(shd.is_holiday_for_zone(d.date(), 'C'))

        # Tournament
        tourn = 0
        for tn_date in df_tn.index:
            if d in pd.date_range(end=tn_date, periods=4):
                tourn = 1

        # Weather
        t = d.isoformat()
        url = f'https://api.darksky.net/forecast/{DARKSKY_API_KEY3}/48.854891,2.203673,{t}?exclude=hourly,currently&flags&lang=fr&units=si'
        response = requests.get(url)
        if response.status_code is not 200:
            print('Http Error')
        weather = response.json()
        temp_max = get_attribute(weather, 'temperatureMax')
        temp_min = get_attribute(weather, 'temperatureMin')
        wind = get_attribute(weather, 'windGust')
        rain = get_attribute(weather, 'precipProbability')

        # add scale rain between 0 and 4 include
        rain = add_scale_rain(rain)

        # Competition a Saint-Cloud
        nbPlay = new_data_df.loc[d, 'nbPlay']

        # Restaurant open ?
        open_res = new_data_df.loc[d, 'open']

        # Number of covers
        nbCov = new_data_df.loc[d, 'nbCov']

        #print(nbCov)
        #print(temp_min)
        #print(type(wind))

        df.loc[d] = pd.Series({
            'fer': fer,
            'vac': vac,
            'tourn': tourn,
            'tempMoy': round((temp_max + temp_min) / 2, 2),
            'tempMax': temp_max,
            'tempMin': temp_min,
            'wind': wind,
            'rain': rain,  ##True if rain > 0.5 else False,
            'nbPlay': nbPlay,
            'open': open_res,
            'nbCov': nbCov
        }, name=d)

    return df


def find_display_cell(ws, name, date):
    # Find column
    done = False
    for c in range(4, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == name:
            done = True
            break

    if not done:
        c, _ = add_display_year(ws)
        if name[-4:] != 'Réel':
            c += 1

    # Find row
    r = date.dayofyear + 1
    if not calendar.isleap(date.year) and r >= 61:
        r += 1

    return r, c


def add_display_year(ws):
    # Get new year and new column id
    # (max_column is not reliable so we have to loop)
    for c in range(5, ws.max_column + 2):
        if ws.cell(row=1, column=c).value is None:
            break
        else:
            last_year = int(ws.cell(row=1, column=c).value[:4])

    # Add Real and Forecast columns
    ws.cell(row=1, column=c).value = str(last_year + 1) + ' Réel'
    ws.cell(row=1, column=c + 1).value = str(last_year + 1) + ' Prédiction'

    return c, last_year + 1


def update_display(pred, df):
    # Download display file from blob storage
    download_blob_from_url(
        blob_url=BLOB_URL,
        output=DISPLAY_FILE,
        overwrite=True,
        credential=BLOB_ACCESS_KEY
    )

    # Load display data
    display_wb = pyxl.load_workbook(DISPLAY_FILE)
    ws = display_wb.active

    ## Update Real column(s)
    # Get the first date missing real data
    # For each Real data column
    year = None
    for c in range(4, ws.max_column + 1):
        if ws.cell(row=1, column=c).value[-4:] == 'Réel':
            # Find the row where we start missing values
            for r in range(2, 368):
                if ws.cell(row=r, column=c).value is None:
                    year = int(ws.cell(row=1, column=c).value[:4])
                    break

            if year is not None:
                break

    # If we didn't find it (all the Real columns are filled-out)
    if year is None:
        # Create new columns
        c, year = add_display_year(ws)
        r = 2

    # Get the corresponding date
    missing_val_date = dt.datetime(year, 1, 1) + dt.timedelta(r - 2)

    # Add missing values
    for d, row in df.loc[missing_val_date:].iterrows():
        if np.isnan(row['nbCov']):
            break
        # Find cell
        r, c = find_display_cell(ws, str(d.year) + ' Réel', d)

        # Put real value in the display dataset
        ws.cell(row=r, column=c).value = row['nbCov']

    ## Update Forecast column(s)
    # Get list of forecast dates
    forecast_ids = list(df.loc[df['nbCov'].isna()].index)

    pred_i = 0
    for d in forecast_ids:
        # Find cell
        r, c = find_display_cell(ws, str(d.year) + ' Prédiction', d)

        # Check if the restaurant is opened
        if not df.loc[d, 'open']:
            ws.cell(row=r, column=c).value = 0
        else:
            ws.cell(row=r, column=c).value = int(round(pred[pred_i, 0]))
            pred_i += 1

    # Verify that we have put all the predictions
    if pred_i < len(pred):
        raise RuntimeError('ERROR: prediction(s) still remaining')

    # Save display file
    display_wb.save(DISPLAY_FILE)

    # Upload display file
    with open(DISPLAY_FILE, 'rb') as data_display_file:
        upload_blob_to_url(
            blob_url=BLOB_URL,
            data=data_display_file,
            overwrite=True,
            credential=BLOB_ACCESS_KEY
        )


def retrieve_file_from_mail():
    """
        Look for a received attachment in the golf's mailbox.
        returns list of the path of the downloaded file(s)
    """
    path_list = []

    # Connect to mailbox
    mail_box = imaplib.IMAP4_SSL('outlook.office365.com', 993)
    mail_box.login(EMAIL_USER, EMAIL_PASS)
    mail_box.select('Inbox')

    # Retrieve mails
    typ, data = mail_box.search(None, 'ALL')

    # For each received mail
    for num in data[0].split():
        print('NEW MAIL\n')
        typ, data = mail_box.fetch(num, '(RFC822)')

        # Get mail content
        raw_email = data[0][1]
        raw_email_string = raw_email.decode('utf-8')
        msg = email.message_from_string(raw_email_string)

        # Search for attachment
        if not msg.is_multipart():
            continue

        for part in msg.walk():
            if part.is_multipart() or part.get('Content-Disposition') is None:
                continue

            # Get file_name
            file_name = part.get_filename()

            if file_name is not None:
                # Decode the name if needed
                if file_name[:2] == '=?':
                    text, enc = email.header.decode_header(file_name)[0]
                    file_name = text.decode(enc)

                # Check that it's an excel file
                if file_name[-5:] != '.xlsx':
                    continue

                # Create the file
                f_str = str(int(time.time())) + str(len(path_list))
                file_path = NEW_DATA_DIR + '/GND_{0}.xlsx'.format(f_str)
                with open(file_path, 'wb') as file:
                    file.write(part.get_payload(decode=True))

                # Add path to the list
                path_list.append(file_path)

        # Delete mail
        mail_box.store(num, '+FLAGS', '\\Deleted')

    # Close mailbox
    mail_box.expunge()
    mail_box.close()
    mail_box.logout()

    return path_list


def check_gnd_files():
    """ Check the number of Golf New Data files and delete some if needed. """
    while len(os.listdir(NEW_DATA_DIR)) > MAX_NB_NEW_DATA_FILES:
        list_of_files = os.listdir(NEW_DATA_DIR)
        full_path = [NEW_DATA_DIR + '/{0}'.format(x) for x in list_of_files]

        oldest_file = min(full_path, key=os.path.getctime)
        print('Deleting', oldest_file)
        os.remove(oldest_file)


def main(data_file, new_data_file):
    print(data_file)
    data_df = pd.read_csv(data_file, parse_dates=True, date_parser=parse_date, index_col=0, sep=';')
    new_data_df = pd.read_csv(
        new_data_file,
        date_parser=parse_date,
        index_col=0,
        header=0,
        sep=';'
        # usecols='B:E'
        # true_values=['oui', 'Oui', 'OUI'],
        # false_values=['non', 'Non', 'NON']
    )
    new_data_df.columns = ['nbPlay', 'open', 'nbCov']
    # Check that 'open' column is boolean
    # if new_data_df.dtypes.open != bool:
    #    raise ValueError('The \'open\' column of the new data should be boolean.')

    # new_data_df = new_data_df.set_index('Date')######
    # new_data_df.index = new_data_df.index.strftime('%d/%m/%Y')######

    data_df = add_rows(data_df, new_data_df)
    # Mise en forme date
    # data_df = pd.to_datetime(data_df.index, dayfirst=True)

    # replace nan by 0
    data_df = data_df.fillna(0)

    check_integrity(data_df)

    # Dirty Code
    data_df = data_df.reset_index()
    # Add for convert date forms
    data_df['Date'] = pd.to_datetime(data_df['Date'], format='%Y-%m-%d').dt.strftime('%d/%m/%Y')
    data_df = data_df.set_index('Date')
    print(data_df.to_string())

    # prepared_df = prepareData(data_df)

    # if prepared_df['nbCov'].isna().any():
    #    # Find the index of the first forecast, and the number of forecasts to perform
    #    forecast_id = prepared_df.loc[prepared_df['nbCov'].isna()].index[0]
    #    forecast_int_id = prepared_df.index.get_loc(forecast_id)
    #    nb_forecast = prepared_df.loc[forecast_id:, 'nbCov'].shape[0]

    data_df.to_csv(ADD_DATA, sep=';', decimal='.')

    # Load the model
    """ with open(TRAIN_DATE_FILE, 'r') as train_date_f:
            date_str = train_date_f.read()
        model = lgb.Booster(model_file=f'gbm_models/gbm{date_str}.txt')

        # Forecast
        print(f'Forecasting from the {forecast_id} for {nb_forecast} days.')
        pred_y = forecast(prepared_df.iloc[forecast_int_id - 7:], model, nb_forecast)

        # Print forecast
        for index, pred in zip(prepared_df.loc[forecast_id:].index, pred_y):
            print(index, pred)

        # Load forecast in the Display dataset
        update_display(pred_y, data_df)

    # Save data file
    data_df.to_csv(DATA_FILE)

    # Check number of new data files
    check_gnd_files()"""


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Video Recognition Bot")
    parser.add_argument("-d", help="file path of the complete dataset",
                        default=DATA_FILE, type=str)
    args = parser.parse_args()
    data_file = args.d

    # new_data_file = NEW_DATA_DIR  # retrieve_file_from_mail()
    print('\n\n\n\n', str(dt.datetime.now()) + ': Executing seq_forecast')

    file = NEW_DATA_FILE
    print('Golf New Data file:', file)
    check_files(data_file, file)

    main(data_file, file)

    """for file in new_data_file:
        print(file)
        print('Golf New Data file:', file)
        check_files(data_file, file)

        main(data_file, file)"""

    """if not new_data_file:
        print('No New Data file.')"""
